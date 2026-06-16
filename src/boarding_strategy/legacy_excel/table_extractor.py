from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import openpyxl

from boarding_strategy.legacy_model.legacy_school_strength import school_strength_from_rank

from .anonymizer import NameAnonymizer
from .formula_inventory import evaluate_simple_chain
from .workbook_inspector import inspect_workbook, write_audit


def import_legacy_workbook(workbook: str | Path, out_dir: str | Path, anonymize: bool = True) -> dict[str, Any]:
    out_dir = Path(out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    wb_formula = openpyxl.load_workbook(workbook, data_only=False, read_only=False)
    wb_values = openpyxl.load_workbook(workbook, data_only=True, read_only=False)
    anonymizer = NameAnonymizer()

    students = extract_stage1_students(wb_values, anonymizer if anonymize else None)
    schools = extract_schools(wb_formula, wb_values)
    toefl = extract_toefl_table(wb_formula)
    gpa = extract_gpa_table(wb_formula)
    history = extract_history(wb_values, anonymizer if anonymize else None)
    rubric = extract_application_rubric(wb_formula, wb_values)

    _write_csv(out_dir / "legacy_stage1_inputs.csv", students)
    _write_csv(out_dir / "legacy_schools.csv", schools)
    _write_csv(out_dir / "toefl_conversion.csv", [{"toefl": k, "eval": v} for k, v in sorted(toefl.items(), reverse=True)])
    _write_csv(out_dir / "gpa_conversion.csv", [{"label": k, "eval": v} for k, v in gpa.items()])
    _write_csv(out_dir / "historical_outcomes.csv", history)
    (out_dir / "application_rubric.json").write_text(json.dumps(rubric, ensure_ascii=False, indent=2), encoding="utf-8")
    if anonymize:
        anonymizer.write_private_map(out_dir.parent / "name_map.csv")
    return {
        "students": len(students),
        "schools": len(schools),
        "toefl_rows": len(toefl),
        "gpa_rows": len(gpa),
        "historical_outcomes": len(history),
        "rubric_categories": len(rubric),
        "out_dir": str(out_dir),
    }


def inspect_and_import(workbook: str | Path, manifest_out: str | Path, extracted_dir: str | Path, audit_out: str | Path) -> dict[str, Any]:
    manifest = inspect_workbook(workbook, manifest_out)
    write_audit(manifest, audit_out)
    summary = import_legacy_workbook(workbook, extracted_dir, anonymize=True)
    manifest["extraction_summary"] = summary
    Path(manifest_out).write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def extract_stage1_students(wb, anonymizer: NameAnonymizer | None = None) -> list[dict[str, Any]]:
    ws = wb["Inboardy - Student Input"]
    headers = [ws.cell(2, c).value for c in range(1, ws.max_column + 1)]
    out = []
    for r in range(3, ws.max_row + 1):
        name = ws.cell(r, 2).value
        if not name:
            continue
        row = {headers[c - 1] or f"col_{c}": ws.cell(r, c).value for c in range(1, min(ws.max_column, 76) + 1)}
        row["student_alias"] = anonymizer.alias(name) if anonymizer else name
        row.pop("학생", None)
        out.append(row)
    return out


def extract_toefl_table(wb) -> dict[int, float]:
    ws = wb["Inboardy Score Backend"]
    values: dict[str, float] = {}
    table: dict[int, float] = {}
    for r in range(13, 134):
        raw_score = ws.cell(r, 22).value
        raw_eval = ws.cell(r, 23).value
        if raw_score is None:
            continue
        if isinstance(raw_eval, (int, float)):
            eval_score = float(raw_eval)
        elif isinstance(raw_eval, str) and raw_eval.startswith("="):
            eval_score = evaluate_simple_chain(raw_eval, values)
        else:
            eval_score = None
        if eval_score is None:
            continue
        score = int(raw_score)
        values[f"W{r}"] = eval_score
        table[score] = round(eval_score, 6)
    return table


def extract_gpa_table(wb) -> dict[str, float]:
    ws = wb["Inboardy Score Backend"]
    out = {}
    for r in range(13, 134):
        label = ws.cell(r, 24).value
        score = ws.cell(r, 25).value
        if label and isinstance(score, (int, float)):
            out[str(label)] = float(score)
    return out


def extract_schools(wb_formula, wb_values) -> list[dict[str, Any]]:
    ws = wb_values["Compatability Backend"]
    schools = []
    for r in range(4, 97):
        name = ws.cell(r, 2).value
        if not name:
            continue
        sports = []
        for c in range(18, min(ws.max_column, 150) + 1):
            if str(ws.cell(r, c).value or "").strip().lower() == "x":
                label = ws.cell(3, c).value
                if label:
                    sports.append(str(label))
        rank = float(ws.cell(r, 1).value or (r - 3))
        schools.append(
            {
                "school_id": f"school_{int(rank):04d}",
                "school_name": name,
                "rank": rank,
                "school_strength": school_strength_from_rank(rank),
                "region": ws.cell(r, 3).value,
                "town_state": ", ".join([str(x) for x in [ws.cell(r, 4).value, ws.cell(r, 5).value] if x]),
                "surroundings": ws.cell(r, 6).value,
                "religion": ws.cell(r, 7).value,
                "student_body_size": _size_band(ws.cell(r, 8).value),
                "boarding_percentage": _boarding_band(ws.cell(r, 10).value),
                "international_student_percentage": _intl_band(ws.cell(r, 12).value),
                "airport_name": ws.cell(r, 14).value,
                "airport_city": ws.cell(r, 15).value,
                "average_travel_time_to_airport_minutes": ws.cell(r, 16).value,
                "sports_offered": "|".join(sorted(set(sports))),
                "source_sheet": "Compatability Backend",
                "source_row": r,
            }
        )
    return schools


def extract_history(wb, anonymizer: NameAnonymizer | None = None) -> list[dict[str, Any]]:
    ws = wb["Misc. Backend Stuff"]
    out = []
    for r in range(2, min(ws.max_row, 1973) + 1):
        cohort, name, school, outcome = ws.cell(r, 1).value, ws.cell(r, 4).value, ws.cell(r, 5).value, ws.cell(r, 6).value
        if not name or not school or not outcome:
            continue
        out.append(
            {
                "anonymized_student_id": anonymizer.alias(name) if anonymizer else name,
                "cohort": cohort,
                "school_name": school,
                "outcome": _normalize_outcome(outcome),
                "legacy_prediction_margin": ws.cell(r, 8).value,
                "school_strength": ws.cell(r, 9).value,
                "source_sheet": "Misc. Backend Stuff",
                "source_row": r,
            }
        )
    return out


def extract_application_rubric(wb_formula, wb_values) -> list[dict[str, Any]]:
    wsf = wb_formula["Application Rubric"]
    wsv = wb_values["Application Rubric"]
    categories = []
    current = None
    for r in range(4, 28):
        cat_label = wsf.cell(r, 1).value
        item_label = wsf.cell(r, 2).value
        if cat_label:
            current = {"name": str(cat_label), "source_row": r, "items": []}
            categories.append(current)
        if current and item_label:
            current["items"].append(
                {
                    "name": str(item_label),
                    "source_cell": f"B{r}",
                    "sample_evidence": wsv.cell(r, 3).value,
                    "sample_score": wsv.cell(r, 5).value,
                    "future_score": wsv.cell(r, 9).value,
                    "formula": wsf.cell(r, 5).value if isinstance(wsf.cell(r, 5).value, str) and wsf.cell(r, 5).value.startswith("=") else None,
                }
            )
    return categories


def _write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8")
        return
    normalized_rows = [{str(k): v for k, v in row.items()} for row in rows]
    keys = sorted({k for row in normalized_rows for k in row})
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=keys)
        writer.writeheader()
        writer.writerows(normalized_rows)


def _size_band(value):
    if value is None:
        return None
    return "Low (less than 450)" if value < 450 else "High (more than 800)" if value > 800 else "Medium (450 - 800)"


def _boarding_band(value):
    if value is None:
        return None
    return "Low (less than 50%)" if value < 0.5 else "All Boarding" if value == 1 else "Medium (50% - 75%)" if value < 0.75 else "High (75% - 99%)"


def _intl_band(value):
    if value is None:
        return None
    return "Low (less than 10%)" if value < 0.1 else "High (more than 20%)" if value > 0.2 else "Average (10 - 20%)"


def _normalize_outcome(value):
    s = str(value or "").strip().lower()
    if s in {"accept", "accepted", "admit", "admitted"}:
        return "admitted"
    if s in {"deny", "denied", "reject", "rejected"}:
        return "denied"
    if "wait" in s:
        return "waitlisted"
    if s in {"w/d", "withdrawn", "withdraw"}:
        return "withdrawn"
    return "unknown"
