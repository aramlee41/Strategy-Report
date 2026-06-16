from __future__ import annotations

import json
from pathlib import Path
from typing import Any

import openpyxl


def inspect_workbook(workbook: str | Path, out: str | Path | None = None) -> dict[str, Any]:
    wb = openpyxl.load_workbook(workbook, data_only=False, read_only=False)
    sheets = []
    for ws in wb.worksheets:
        formula_samples = []
        nonempty = 0
        used_rows = []
        likely_headers = []
        for row in ws.iter_rows():
            values = [cell.value for cell in row]
            if any(v is not None for v in values):
                used_rows.append(row[0].row)
                nonempty += sum(v is not None for v in values)
                text_count = sum(isinstance(v, str) and not v.startswith("=") for v in values[:30])
                if text_count >= 3 and len(likely_headers) < 5:
                    likely_headers.append(row[0].row)
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("=") and len(formula_samples) < 12:
                    formula_samples.append({"cell": cell.coordinate, "formula": cell.value[:240]})
        sheets.append(
            {
                "sheet": ws.title,
                "max_row": ws.max_row,
                "max_col": ws.max_column,
                "used_row_min": min(used_rows) if used_rows else None,
                "used_row_max": max(used_rows) if used_rows else None,
                "nonempty_cells": nonempty,
                "formula_count": sum(1 for row in ws.iter_rows() for c in row if isinstance(c.value, str) and c.value.startswith("=")),
                "likely_header_rows": likely_headers,
                "formula_samples": formula_samples,
            }
        )
    manifest = {
        "workbook_name": Path(workbook).name,
        "sheet_count": len(wb.sheetnames),
        "sheets": sheets,
        "private_data_warning": "Workbook may contain real student names. Extracted data belongs in data/private.",
    }
    if out:
        out = Path(out)
        out.parent.mkdir(parents=True, exist_ok=True)
        out.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")
    return manifest


def write_audit(manifest: dict[str, Any], out: str | Path) -> Path:
    path = Path(out)
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = ["# Migration Audit", "", f"Workbook: `{manifest['workbook_name']}`", f"Sheets: {manifest['sheet_count']}", ""]
    for sheet in manifest["sheets"]:
        lines.append(f"## {sheet['sheet']}")
        lines.append(f"- Used rows: {sheet['used_row_min']} to {sheet['used_row_max']}")
        lines.append(f"- Max columns: {sheet['max_col']}")
        lines.append(f"- Formula count: {sheet['formula_count']}")
        lines.append(f"- Likely header rows: {sheet['likely_header_rows']}")
        if sheet["formula_samples"]:
            lines.append("- Formula sample: `" + sheet["formula_samples"][0]["formula"].replace("`", "'") + "`")
        lines.append("")
    lines.append("## Privacy")
    lines.append("Private extraction outputs are intentionally ignored by git.")
    path.write_text("\n".join(lines), encoding="utf-8")
    return path
